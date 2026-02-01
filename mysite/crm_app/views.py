from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.views import APIView

from .serializers import *
from .permissions import *
from .filters import AppointmentFilter
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .serializers import VerifyResetCodeSerializer
from django.db.models import Sum, F, DecimalField, ExpressionWrapper
from django.db.models import Count, Q
from datetime import timedelta, date
import openpyxl
from django.http import HttpResponse
from django.db.models.functions import TruncDate




class CustomLoginView(generics.GenericAPIView):
    serializer_class = CustomLoginSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class LogoutView(generics.GenericAPIView):
    serializer_class = LogoutSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            refresh_token = serializer.validated_data['refresh']
            token = RefreshToken(refresh_token)
            token.blacklist()
            return Response(status=status.HTTP_205_RESET_CONTENT)
        except Exception:
            return Response({'detail': 'Невалидный токен'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def verify_reset_code(request):
    serializer = VerifyResetCodeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Пароль успешно сброшен.'}, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


"""Admin"""
class AdminUserCreateView(generics.CreateAPIView):
    serializer_class = AdminUserCreateSerializer
    permission_classes = [IsAuthenticated, IsAdmin]


class AdminAppointmentListAPIView(generics.ListAPIView):
    serializer_class = AdminAppointmentListSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    filter_backends = [DjangoFilterBackend]

    filterset_fields = {
        "doctor": ["exact"],
        "start_time": ["date__exact", "date__gte", "date__lte"],
    }

    def get_queryset(self):
        return (
            Appointment.objects
            .select_related("patient", "doctor__user")
            .prefetch_related("payments")
            .order_by("-start_time")
        )

class AdminAddPatientAPIView(generics.CreateAPIView):
    serializer_class = AdminAddPatientSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        patient = serializer.save()

        return Response(
            {
                "message": "Пациент успешно добавлен",
                "patient_id": patient.id
            },
            status=status.HTTP_201_CREATED
        )

# views/admin_edit_appointment.py
class AdminAppointmentEditAPIView(generics.RetrieveUpdateAPIView):
    queryset = Appointment.objects.select_related(
        "patient", "doctor__user", "department", "service", "registrar"
    )
    serializer_class = AdminAppointmentEditSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

class AdminPatientAppointmentHistoryAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request, patient_id):
        qs = (
            Appointment.objects
            .filter(patient_id=patient_id)
            .select_related(
                "registrar",
                "department",
                "doctor__user",
                "service"
            )
            .order_by("-created_at")
        )

        # 🔹 статистика (то, что ты просил)
        stats = {
            "total": qs.count(),
            "queue": qs.filter(status="queue").count(),
            "completed": qs.filter(status="completed").count(),
            "cancelled": qs.filter(status="cancelled").count(),
        }

        data = {
            "stats": stats,
            "results": AdminPatientAppointmentHistorySerializer(qs, many=True).data
        }

        return Response(data)

class AdminAppointmentDeleteAPIView(generics.DestroyAPIView):
    queryset = Appointment.objects.all()
    permission_classes = [IsAuthenticated, IsAdmin]

class AdminPatientVisitHistoryAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request, patient_id):
        qs = (
            Appointment.objects
            .filter(
                patient_id=patient_id,
                status="completed"
            )
            .select_related(
                "registrar",
                "department",
                "doctor__user",
                "service"
            )
            .order_by("-created_at")
        )

        # 🔹 общая статистика (как на макете)
        all_qs = Appointment.objects.filter(patient_id=patient_id)

        stats = {
            "total": all_qs.filter(status="completed").count(),
            "queue": all_qs.filter(status="queue").count(),
            "completed": all_qs.filter(status="completed").count(),
            "cancelled": all_qs.filter(status="cancelled").count(),
        }

        return Response({
            "stats": stats,
            "results": AdminPatientVisitHistorySerializer(qs, many=True).data
        })


class AdminPatientPaymentAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request, patient_id):
        qs = (
            Payment.objects
            .filter(appointment__patient_id=patient_id)
            .select_related(
                "appointment__department",
                "appointment__doctor__user",
                "appointment__service",
            )
            .order_by("-created_at")
        )

        total = qs.aggregate(total=Sum("amount"))["total"] or 0
        cash = qs.filter(method="cash").aggregate(
            total=Sum("amount")
        )["total"] or 0
        card = qs.filter(method="card").aggregate(
            total=Sum("amount")
        )["total"] or 0

        return Response({
            "summary": {
                "total": total,
                "cash": cash,
                "card": card,
            },
            "results": AdminPatientPaymentSerializer(qs, many=True).data
        })

class AdminPatientDetailAPIView(generics.RetrieveAPIView):
    queryset = Patient.objects.all()
    serializer_class = AdminPatientDetailSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    lookup_field = "id"

class AdminAppointmentPaymentCreateAPIView(generics.CreateAPIView):
    serializer_class = AdminAppointmentPaymentCreateSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        payment = serializer.save()

        return Response(
            {
                "message": "Оплата успешно принята",
                "payment_id": payment.id
            },
            status=status.HTTP_201_CREATED
        )

from rest_framework.filters import SearchFilter


class AdminDoctorListAPIView(generics.ListAPIView):
    queryset = Doctor.objects.select_related("user", "department")
    serializer_class = AdminDoctorListSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    filter_backends = [SearchFilter, DjangoFilterBackend]
    search_fields = [
        "user__first_name",
        "user__last_name",
        "cabinet",
    ]
    filterset_fields = ["department"]

class AdminDoctorCreateAPIView(generics.CreateAPIView):
    serializer_class = AdminDoctorCreateUpdateSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

class AdminDoctorDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Doctor.objects.select_related("user", "department")
    serializer_class = AdminDoctorCreateUpdateSerializer
    permission_classes = [IsAuthenticated, IsAdmin]


class AdminAnalyticsAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        period = request.query_params.get("period", "week")

        today = date.today()

        if period == "month":
            start_date = today - timedelta(days=30)
        elif period == "day":
            start_date = today - timedelta(days=1)
        else:  # week
            start_date = today - timedelta(days=7)

        # ===== ВРАЧИ =====
        doctors_count = Doctor.objects.count()

        # ===== ЗАПИСИ =====
        qs = Appointment.objects.filter(
            created_at__date__gte=start_date
        )

        total_appointments = qs.count()
        cancelled = qs.filter(status="cancelled").count()

        # рост / падение (примитивно, но честно)
        growth_percent = round(
            (total_appointments / max(1, cancelled)) * 10, 1
        )
        decline_percent = round(
            (cancelled / max(1, total_appointments)) * 10, 1
        )

        # ===== КЛИЕНТЫ =====
        total_patients = qs.values("patient").distinct().count()

        primary = (
            Patient.objects
            .annotate(cnt=Count("appointments"))
            .filter(cnt=1)
            .count()
        )
        repeat = max(0, total_patients - primary)

        primary_percent = int((primary / max(1, total_patients)) * 100)
        repeat_percent = 100 - primary_percent

        # ===== ГРАФИК =====
        chart = (
            qs.values("created_at__date")
            .annotate(
                total=Count("id"),
                cancelled=Count("id", filter=Q(status="cancelled"))
            )
            .order_by("created_at__date")
        )

        chart_data = []
        for item in chart:
            chart_data.append({
                "date": item["created_at__date"],
                "total": item["total"],
                "cancelled": qs.filter(
                    created_at__date=item["created_at__date"],
                    status="cancelled"
                ).count()
            })

        return Response({
            "growth_percent": growth_percent,
            "decline_percent": -decline_percent,
            "doctors_count": doctors_count,

            "total_patients": total_patients,
            "primary_percent": primary_percent,
            "repeat_percent": repeat_percent,

            "chart": chart_data
        })

class AdminDetailedReportAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        qs = Payment.objects.select_related(
            "appointment__patient",
            "appointment__service",
            "appointment__department",
            "appointment__doctor__user",
        )

        # ===== ФИЛЬТРЫ =====
        doctor_id = request.query_params.get("doctor")
        department_id = request.query_params.get("department")
        search = request.query_params.get("search")
        period = request.query_params.get("period")  # day | week | month
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        if doctor_id:
            qs = qs.filter(appointment__doctor_id=doctor_id)

        if department_id:
            qs = qs.filter(appointment__department_id=department_id)

        if search:
            qs = qs.filter(
                appointment__patient__full_name__icontains=search
            )

        # ===== ПЕРИОД =====
        today = date.today()

        if period == "day":
            qs = qs.filter(created_at__date=today)

        elif period == "week":
            qs = qs.filter(created_at__date__gte=today - timedelta(days=7))

        elif period == "month":
            qs = qs.filter(created_at__date__gte=today - timedelta(days=30))

        # ===== КАЛЕНДАРЬ =====
        if date_from and date_to:
            qs = qs.filter(
                created_at__date__gte=date_from,
                created_at__date__lte=date_to
            )

        # ===== ИТОГИ =====
        total_count = qs.count()
        total_sum = qs.aggregate(total=Sum("amount"))["total"] or 0
        cash_sum = qs.filter(method="cash").aggregate(
            s=Sum("amount")
        )["s"] or 0
        card_sum = qs.filter(method="card").aggregate(
            s=Sum("amount")
        )["s"] or 0

        return Response({
            "results": AdminDetailedReportRowSerializer(qs, many=True).data,
            "summary": {
                "total_count": total_count,
                "total_sum": total_sum,
                "cash": cash_sum,
                "card": card_sum,
            }
        })

class AdminDetailedReportExcelAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        qs = Payment.objects.select_related(
            "appointment__patient",
            "appointment__service",
            "appointment__doctor__user",
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        ws.append([
            "Дата", "Пациент", "Услуга",
            "Тип оплаты", "Сумма", "Врач"
        ])

        for p in qs:
            ws.append([
                p.created_at.strftime("%d.%m.%Y"),
                p.appointment.patient.full_name,
                p.appointment.service.name,
                "Наличные" if p.method == "cash" else "Безналичные",
                float(p.amount),
                p.appointment.doctor.user.get_full_name(),
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="report.xlsx"'
        wb.save(response)

        return response

class AdminDoctorCloseReportAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        doctor_id = request.query_params.get("doctor")
        period = request.query_params.get("period", "month")

        qs = Payment.objects.filter(
            appointment__status="completed"
        )

        if doctor_id:
            qs = qs.filter(
                appointment__doctor_id=doctor_id
            )

        # ===== ПЕРИОД =====
        today = date.today()

        if period == "day":
            qs = qs.filter(created_at__date=today)

        elif period == "month":
            qs = qs.filter(
                created_at__year=today.year,
                created_at__month=today.month
            )

        # ===== ГРУППИРОВКА =====
        rows = (
            qs
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(total_sum=Sum("amount"))
            .order_by("day")
        )

        data = []
        total = 0

        for idx, row in enumerate(rows, start=1):
            total += row["total_sum"]
            data.append({
                "id": idx,
                "date": row["day"],
                "total_sum": row["total_sum"]
            })

        return Response({
            "doctor": doctor_id,
            "results": data,
            "total_sum": total
        })

class AdminDoctorCloseReportExcelAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        doctor_id = request.query_params.get("doctor")

        qs = Payment.objects.filter(
            appointment__status="completed"
        )

        if doctor_id:
            qs = qs.filter(
                appointment__doctor_id=doctor_id
            )

        rows = (
            qs
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(total_sum=Sum("amount"))
            .order_by("day")
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Закрытия врача"

        ws.append(["№", "Дата", "Сумма"])

        total = 0
        for idx, row in enumerate(rows, start=1):
            ws.append([
                idx,
                row["day"].strftime("%d.%m.%Y"),
                float(row["total_sum"])
            ])
            total += row["total_sum"]

        ws.append(["", "Итого:", float(total)])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="doctor_close_report.xlsx"'
        )

        wb.save(response)
        return response

class AdminSummaryReportAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        qs = Payment.objects.filter(
            appointment__status="completed"
        )

        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)

        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # ===== ОБЩИЕ СУММЫ =====
        total_cash = qs.filter(method="cash").aggregate(
            s=Sum("amount")
        )["s"] or 0

        total_card = qs.filter(method="card").aggregate(
            s=Sum("amount")
        )["s"] or 0

        total_sum = total_cash + total_card

        # ===== БОНУС ВРАЧЕЙ =====
        doctor_bonus_expr = ExpressionWrapper(
            F("amount") * F("appointment__doctor__bonus_percent") / 100,
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )

        doctors_cash = qs.filter(method="cash").aggregate(
            s=Sum(doctor_bonus_expr)
        )["s"] or 0

        doctors_card = qs.filter(method="card").aggregate(
            s=Sum(doctor_bonus_expr)
        )["s"] or 0

        doctors_total = doctors_cash + doctors_card

        # ===== КЛИНИКА =====
        clinic_cash = total_cash - doctors_cash
        clinic_card = total_card - doctors_card

        return Response({
            "total_cash": total_cash,
            "total_card": total_card,
            "total_sum": total_sum,

            "doctors_total": doctors_total,
            "doctors_cash": doctors_cash,
            "doctors_card": doctors_card,

            "clinic_cash": clinic_cash,
            "clinic_card": clinic_card,
        })

class AdminSummaryReportExcelAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        api = AdminSummaryReportAPIView()
        api.request = request
        data = api.get(request).data

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Сводный отчет"

        ws.append(["Показатель", "Сумма"])

        ws.append(["Оплачено наличными", data["total_cash"]])
        ws.append(["Оплачено безналичными", data["total_card"]])
        ws.append(["Общая сумма", data["total_sum"]])

        ws.append([])
        ws.append(["Врачам всего", data["doctors_total"]])
        ws.append(["Врачам наличными", data["doctors_cash"]])
        ws.append(["Врачам безналичными", data["doctors_card"]])

        ws.append([])
        ws.append(["Клиника наличными", data["clinic_cash"]])
        ws.append(["Клиника безналичными", data["clinic_card"]])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="summary_report.xlsx"'
        )

        wb.save(response)
        return response

class AdminCalendarListAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        start = request.query_params.get("start")
        end = request.query_params.get("end")
        doctor_id = request.query_params.get("doctor")
        department_id = request.query_params.get("department")

        qs = Appointment.objects.select_related(
            "patient",
            "doctor__user",
            "service",
            "department",
        )

        if start and end:
            qs = qs.filter(
                start_time__date__gte=start,
                start_time__date__lte=end
            )

        if doctor_id:
            qs = qs.filter(doctor_id=doctor_id)

        if department_id:
            qs = qs.filter(department_id=department_id)

        serializer = AdminCalendarAppointmentSerializer(qs, many=True)
        return Response(serializer.data)

class AdminCalendarCreateAPIView(generics.CreateAPIView):
    serializer_class = AdminCalendarCreateSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def perform_create(self, serializer):
        serializer.save(registrar=self.request.user)

class AdminCalendarUpdateAPIView(generics.UpdateAPIView):
    queryset = Appointment.objects.all()
    serializer_class = AdminCalendarCreateSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

class AdminCalendarDeleteAPIView(generics.DestroyAPIView):
    queryset = Appointment.objects.all()
    permission_classes = [IsAuthenticated, IsAdmin]

# ===== PRICE LIST =====

class AdminPriceListAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        qs = (
            Department.objects
            .prefetch_related("services")
            .order_by("name")
        )

        return Response(
            AdminPriceListDepartmentSerializer(qs, many=True).data
        )


# ===== SERVICE CRUD =====

class AdminServiceCreateAPIView(generics.CreateAPIView):
    serializer_class = AdminServiceCreateSerializer
    permission_classes = [IsAuthenticated, IsAdmin]


class AdminServiceUpdateAPIView(generics.UpdateAPIView):
    queryset = Service.objects.all()
    serializer_class = AdminServiceCreateSerializer
    permission_classes = [IsAuthenticated, IsAdmin]


class AdminServiceDeleteAPIView(generics.DestroyAPIView):
    queryset = Service.objects.all()
    permission_classes = [IsAuthenticated, IsAdmin]

"""Receptionist"""
# ✅ ИСПРАВЛЕНО: Убрал фильтр по registrar
class ReceptionistAppointmentListAPIView(generics.ListAPIView):
    serializer_class = ReceptionistAppointmentListSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]
    filter_backends = [DjangoFilterBackend]

    filterset_fields = {
        "doctor": ["exact"],
        "start_time": ["date__exact", "date__gte", "date__lte"],
    }

    def get_queryset(self):
        # ✅ Показываем ВСЕ записи, а не только созданные этим регистратором
        return (
            Appointment.objects
            .select_related("patient", "doctor__user")
            .prefetch_related("payments")
            .order_by("-start_time")
        )

class ReceptionistAddPatientAPIView(generics.CreateAPIView):
    serializer_class = ReceptionistAddPatientSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(
            data=request.data,
            context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        patient = serializer.save()

        return Response(
            {
                "message": "Пациент успешно добавлен",
                "patient_id": patient.id
            },
            status=status.HTTP_201_CREATED
        )

# ✅ ИСПРАВЛЕНО: Убрал фильтр по registrar
class ReceptionistAppointmentEditAPIView(generics.RetrieveUpdateAPIView):
    serializer_class = ReceptionistAppointmentEditSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]

    def get_queryset(self):
        # ✅ Разрешаем редактировать ВСЕ записи
        return Appointment.objects.all()


# ✅ ИСПРАВЛЕНО: Убрал фильтр по registrar
class ReceptionistPatientAppointmentHistoryAPIView(generics.ListAPIView):
    serializer_class = ReceptionistPatientAppointmentHistorySerializer
    permission_classes = [IsAuthenticated, IsReceptionist]

    def get_queryset(self):
        # ✅ Показываем ВСЕ записи пациента
        return Appointment.objects.filter(
            patient_id=self.kwargs["patient_id"]
        ).order_by("-created_at")


# ✅ ИСПРАВЛЕНО: Убрал фильтр по registrar
class ReceptionistPatientPaymentAPIView(generics.ListAPIView):
    serializer_class = ReceptionistPatientPaymentSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]

    def get_queryset(self):
        # ✅ Показываем ВСЕ платежи пациента
        return Payment.objects.filter(
            appointment__patient_id=self.kwargs["patient_id"]
        ).order_by("-created_at")


class ReceptionistAppointmentPaymentCreateAPIView(generics.CreateAPIView):
    serializer_class = ReceptionistAppointmentPaymentCreateSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]

    def get_serializer_context(self):
        return {"request": self.request}

# ===== RECEPTIONIST PROFILE =====
class ReceptionistProfileAPIView(APIView):
    permission_classes = [IsAuthenticated, IsReceptionist]

    def get(self, request):
        serializer = ReceptionistProfileSerializer(request.user)
        return Response(serializer.data)

# ===== LIST =====
class ReceptionistDoctorListAPIView(generics.ListAPIView):
    queryset = Doctor.objects.select_related("user", "department")
    serializer_class = ReceptionistDoctorListSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]

    filter_backends = [SearchFilter, DjangoFilterBackend]
    search_fields = [
        "user__first_name",
        "user__last_name",
        "cabinet",
    ]
    filterset_fields = ["department"]


# ===== CREATE =====
class ReceptionistDoctorCreateAPIView(generics.CreateAPIView):
    serializer_class = ReceptionistDoctorCreateUpdateSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]


# ===== DETAIL / UPDATE / DELETE =====
class ReceptionistDoctorDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Doctor.objects.select_related("user", "department")
    serializer_class = ReceptionistDoctorCreateUpdateSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]

class ReceptionistDoctorListAPIView(generics.ListAPIView):
    queryset = Doctor.objects.select_related("user", "department")
    serializer_class = ReceptionistDoctorListSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]

    filter_backends = [SearchFilter, DjangoFilterBackend]
    search_fields = ["user__first_name", "user__last_name", "cabinet"]
    filterset_fields = ["department"]


class ReceptionistPriceListAPIView(APIView):
    permission_classes = [IsAuthenticated, IsReceptionist]

    def get(self, request):
        qs = Department.objects.prefetch_related("services")
        return Response(
            ReceptionistPriceListDepartmentSerializer(qs, many=True).data
        )

class ReceptionistDetailedReportAPIView(
    AdminDetailedReportAPIView
):
    permission_classes = [IsAuthenticated, IsReceptionist]

class ReceptionistSummaryReportAPIView(
    AdminSummaryReportAPIView
):
    permission_classes = [IsAuthenticated, IsReceptionist]


class ReceptionistCalendarListAPIView(
    AdminCalendarListAPIView
):
    permission_classes = [IsAuthenticated, IsReceptionist]

class ReceptionistCalendarCreateAPIView(generics.CreateAPIView):
    serializer_class = ReceptionistCalendarCreateSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]

    def perform_create(self, serializer):
        serializer.save(registrar=self.request.user)

class ReceptionistCalendarUpdateAPIView(generics.UpdateAPIView):
    queryset = Appointment.objects.all()
    serializer_class = ReceptionistCalendarCreateSerializer
    permission_classes = [IsAuthenticated, IsReceptionist]

class ReceptionistCalendarDeleteAPIView(generics.DestroyAPIView):
    queryset = Appointment.objects.all()
    permission_classes = [IsAuthenticated, IsReceptionist]


"""Doctor"""


class DoctorCalendarAPIView(generics.ListAPIView):
    serializer_class = DoctorCalendarSerializer
    permission_classes = [IsAuthenticated, IsDoctor]

    def get_queryset(self):
        return Appointment.objects.filter(
            doctor__user=self.request.user
        ).select_related(
            "patient", "doctor__user", "service", "department"
        ).prefetch_related("payments")


# ✅ ПРАВИЛЬНО
class DoctorAppointmentUpdateAPIView(generics.UpdateAPIView):
    serializer_class = DoctorAppointmentUpdateSerializer
    permission_classes = [IsAuthenticated, IsDoctor]

    def get_queryset(self):
        return Appointment.objects.filter(
            doctor__user=self.request.user
        )


# ✅ ПРАВИЛЬНО
class DoctorProfileAPIView(generics.RetrieveUpdateAPIView):
    serializer_class = DoctorProfileSerializer
    permission_classes = [IsAuthenticated, IsDoctor]

    def get_object(self):
        return Doctor.objects.get(user=self.request.user)


# ✅ ПРАВИЛЬНО
class DoctorPatientAppointmentsAPIView(generics.ListAPIView):
    serializer_class = DoctorPatientAppointmentSerializer
    permission_classes = [IsAuthenticated, IsDoctor]

    def get_queryset(self):
        return Appointment.objects.filter(
            patient_id=self.kwargs["patient_id"],
            doctor__user=self.request.user
        ).select_related("service", "department")


# ✅ ПРАВИЛЬНО
class DoctorPatientPaymentsAPIView(generics.ListAPIView):
    serializer_class = DoctorPatientPaymentSerializer
    permission_classes = [IsAuthenticated, IsDoctor]

    def get_queryset(self):
        return Payment.objects.filter(
            appointment__patient_id=self.kwargs["patient_id"],
            appointment__doctor__user=self.request.user
        ).select_related("appointment__service")


# ✅ ИСПРАВЛЕНО: Теперь врач видит только своих пациентов
class DoctorPatientDetailAPIView(generics.RetrieveAPIView):
    serializer_class = DoctorPatientDetailSerializer
    permission_classes = [IsAuthenticated, IsDoctor]

    def get_queryset(self):
        # ✅ Только пациенты, которые были на приёме у этого врача
        return Patient.objects.filter(
            appointments__doctor__user=self.request.user
        ).distinct()


# ✅ ПРАВИЛЬНО
class DoctorNotificationListAPIView(APIView):
    permission_classes = [IsAuthenticated, IsDoctor]

    def get(self, request):
        qs = Notification.objects.filter(
            recipient=request.user
        ).select_related("appointment__patient", "appointment__department")

        return Response({
            "unread_count": qs.filter(is_read=False).count(),
            "results": DoctorNotificationSerializer(qs, many=True).data
        })


# ✅ ИСПРАВЛЕНО: Добавлена обработка ошибок
class DoctorNotificationReadAPIView(APIView):
    permission_classes = [IsAuthenticated, IsDoctor]

    def post(self, request, pk):
        try:
            notification = Notification.objects.get(
                pk=pk,
                recipient=request.user
            )
            notification.is_read = True
            notification.save(update_fields=["is_read"])

            return Response({"status": "ok"})

        except Notification.DoesNotExist:
            return Response(
                {"error": "Уведомление не найдено"},
                status=status.HTTP_404_NOT_FOUND
            )